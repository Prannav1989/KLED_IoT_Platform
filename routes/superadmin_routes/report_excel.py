import io
import csv
from collections import defaultdict
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from models import SensorData, Parameter, Company
from extensions import db


class ExcelReportGenerator:
    """
    Excel / CSV generator for historical sensor data

    Layout:
    ----------------------------------------------------
    | Company Name (Merged Header)                      |
    ----------------------------------------------------
    | Timestamp | Sensor1-Temp | Sensor1-Hum | Sensor2 |
    ----------------------------------------------------
    | Min Values |    value1   |    value2   |  value3 |
    ----------------------------------------------------
    | Max Values |    value1   |    value2   |  value3 |
    ----------------------------------------------------
    
    New Feature: Time Interval Based Averaging
    ------------------------------------------
    Supports time intervals: 1 minute, 5 minutes, 1 hour, 6 hours, 1 day
    When multiple readings exist in an interval, calculates average values
    """

    def __init__(self):
        # Light pastel color palette for professional reports
        self.COLORS = {
            'company_header': 'FFE3F2FD',  # Very light blue
            'sensor_header': 'FFF3E5F5',   # Very light purple
            'param_header': 'FFE8F5E8',    # Very light green
            'min_row': 'FFE8F5E9',         # Light green for min row
            'max_row': 'FFFCE4EC',         # Light pink for max row
            'border': 'FFBDBDBD',          # Light gray border
        }
        
        # Time intervals in minutes
        self.TIME_INTERVALS = {
            '1 minute': 1,
            '5 minutes': 5,
            '1 hour': 60,
            '6 hours': 360,
            '1 day': 1440
        }

    def generate_device_report(
        self,
        devices,
        company_id,
        start_date=None,
        end_date=None,
        file_type="excel",  # excel | csv
        time_interval=None  # New parameter: time interval for averaging
    ):
        # Validate time interval if provided
        if time_interval and time_interval not in self.TIME_INTERVALS:
            raise ValueError(f"Invalid time interval. Must be one of: {list(self.TIME_INTERVALS.keys())}")
            
        if file_type == "csv":
            return self._generate_csv(devices, company_id, start_date, end_date, time_interval)

        return self._generate_excel(devices, company_id, start_date, end_date, time_interval)

    # ==========================================================
    # HELPER FUNCTIONS FOR TIME INTERVAL AVERAGING
    # ==========================================================
    def _get_interval_start_time(self, timestamp, interval_minutes):
        """
        Calculate the start time of the interval for a given timestamp.
        
        Args:
            timestamp: datetime object
            interval_minutes: interval length in minutes
            
        Returns:
            datetime: Start time of the interval
        """
        if interval_minutes == 1:
            # 1 minute interval: round to minute
            return timestamp.replace(second=0, microsecond=0)
        elif interval_minutes == 5:
            # 5 minute interval: round down to nearest 5 minutes
            minute = (timestamp.minute // 5) * 5
            return timestamp.replace(minute=minute, second=0, microsecond=0)
        elif interval_minutes == 60:
            # 1 hour interval: round to hour
            return timestamp.replace(minute=0, second=0, microsecond=0)
        elif interval_minutes == 360:
            # 6 hour interval: round down to nearest 6 hours
            hour = (timestamp.hour // 6) * 6
            return timestamp.replace(hour=hour, minute=0, second=0, microsecond=0)
        elif interval_minutes == 1440:
            # 1 day interval: round to day
            return timestamp.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            # Default: round to minute
            return timestamp.replace(second=0, microsecond=0)

    def _fetch_averaged_data(self, devices, start_date, end_date, interval_minutes):
        """
        Fetch raw data and perform averaging in Python.
        Database-agnostic: works with SQLite, PostgreSQL, etc.
        """

        device_ids = [d.id for d in devices]

        query = SensorData.query.filter(
            SensorData.device_id.in_(device_ids)
        )

        if start_date:
            query = query.filter(SensorData.timestamp >= start_date)
        if end_date:
            query = query.filter(SensorData.timestamp <= end_date)

        query = query.order_by(SensorData.timestamp.asc())
        records = query.all()

        # ✅ CORRECT STRUCTURE
        grouped_data = defaultdict(list)

        for record in records:
            if record.value is None:
                continue

            interval_start = self._get_interval_start_time(
                record.timestamp,
                interval_minutes
            )

            key = (record.device_id, record.parameter_id, interval_start)

            grouped_data[key].append(record.value)

        averaged_records = []
        for (device_id, parameter_id, interval_start), values in grouped_data.items():
            if values:
                avg_value = sum(values) / len(values)

                averaged_records.append({
                    'device_id': device_id,
                    'parameter_id': parameter_id,
                    'avg_value': avg_value,
                    'interval_start': interval_start
                })

        averaged_records.sort(
            key=lambda x: (x['interval_start'], x['device_id'], x['parameter_id'])
        )

        return averaged_records


    def _fetch_raw_data(self, devices, start_date, end_date):
        """
        Fetch raw sensor data (original behavior).
        """
        query = SensorData.query.filter(
            SensorData.device_id.in_([d.id for d in devices])
        )

        if start_date:
            query = query.filter(SensorData.timestamp >= start_date)
        if end_date:
            query = query.filter(SensorData.timestamp <= end_date)

        query = query.order_by(SensorData.timestamp.asc())
        return query.all()

    # ==========================================================
    # EXCEL EXPORT (UPDATED WITH TIME INTERVAL SUPPORT)
    # ==========================================================
    def _generate_excel(self, devices, company_id, start_date, end_date, time_interval=None):
        wb = Workbook()
        ws = wb.active
        
        # Set worksheet title based on time interval
        if time_interval:
            ws.title = f"Sensor Data ({time_interval} avg)"
        else:
            ws.title = "Sensor Historical Data"

        company = Company.query.get(company_id)
        company_name = company.name if company else "Company Report"
        
        # Add time interval info to header if applicable
        if time_interval:
            company_name = f"{company_name} - {time_interval} Average"

        # ---------- STYLES ----------
        border_side = Side(border_style="thin", color=self.COLORS['border'])
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        company_fill = PatternFill(start_color=self.COLORS['company_header'],
                                  end_color=self.COLORS['company_header'],
                                  fill_type="solid")
        
        sensor_fill = PatternFill(start_color=self.COLORS['sensor_header'],
                                 end_color=self.COLORS['sensor_header'],
                                 fill_type="solid")
        
        param_fill = PatternFill(start_color=self.COLORS['param_header'],
                                end_color=self.COLORS['param_header'],
                                fill_type="solid")
        
        min_fill = PatternFill(start_color=self.COLORS['min_row'],
                              end_color=self.COLORS['min_row'],
                              fill_type="solid")
        
        max_fill = PatternFill(start_color=self.COLORS['max_row'],
                              end_color=self.COLORS['max_row'],
                              fill_type="solid")
        
        header_font = Font(bold=True)
        min_max_font = Font(bold=True, color="FF1A237E")  # Dark blue for contrast
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # ---------- FETCH SENSOR DATA (WITH OR WITHOUT AVERAGING) ----------
        if time_interval:
            # Use averaged data
            interval_minutes = self.TIME_INTERVALS[time_interval]
            averaged_records = self._fetch_averaged_data(devices, start_date, end_date, interval_minutes)
            
            # Convert averaged records to format compatible with existing code
            records = []
            for record in averaged_records:
                # Create a mock SensorData-like object
                class MockRecord:
                    def __init__(self, device_id, parameter_id, avg_value, interval_start):
                        self.device_id = device_id
                        self.parameter_id = parameter_id
                        self.value = avg_value
                        self.timestamp = interval_start
                
                records.append(MockRecord(
                    record['device_id'],
                    record['parameter_id'],
                    record['avg_value'],
                    record['interval_start']
                ))
        else:
            # Use raw data (original behavior)
            records = self._fetch_raw_data(devices, start_date, end_date)

        # ---------- PARAMETER MAP & HEADERS ----------
        column_map = {}
        current_col = 2  # Column B (A is timestamp)
        param_details = []

        # Row 1: Company Header (will be merged dynamically later)
        ws["A1"] = company_name
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].fill = company_fill
        ws["A1"].alignment = center_alignment
        ws["A1"].border = border

        # Row 2: Sensor Names
        # Row 3: Parameter Headers
        ws["A3"] = "Timestamp"
        ws["A3"].font = header_font
        ws["A3"].fill = param_fill
        ws["A3"].alignment = center_alignment
        ws["A3"].border = border

        # Add time interval info in row 4 if applicable
        if time_interval:
            ws["A4"] = f"Data averaged by {time_interval} intervals"
            ws["A4"].font = Font(italic=True)
            ws["A4"].alignment = Alignment(horizontal="left")
            data_start_row = 5  # Start data from row 5
        else:
            data_start_row = 4  # Original: start data from row 4

        # Collect parameter information and setup headers
        for device in devices:
            parameters = Parameter.query.filter_by(device_id=device.id).all()
            if not parameters:
                continue

            start_col = current_col
            end_col = start_col + len(parameters) - 1

            # Merge cells for sensor name
            ws.merge_cells(
                start_row=2,
                start_column=start_col,
                end_row=2,
                end_column=end_col
            )
            
            sensor_cell = ws.cell(row=2, column=start_col)
            sensor_cell.value = device.name
            sensor_cell.font = header_font
            sensor_cell.fill = sensor_fill
            sensor_cell.alignment = center_alignment
            sensor_cell.border = border

            # Parameter headers
            for param in parameters:
                header = f"{param.name}"
                if param.unit:
                    header += f" ({param.unit})"

                param_cell = ws.cell(row=3, column=current_col)
                param_cell.value = header
                param_cell.font = header_font
                param_cell.fill = param_fill
                param_cell.alignment = center_alignment
                param_cell.border = border

                # Store parameter details
                param_details.append({
                    'device_id': device.id,
                    'parameter_id': param.id,
                    'device_name': device.name,
                    'param_name': param.name,
                    'unit': param.unit,
                    'column': current_col
                })
                
                column_map[(device.id, param.id)] = current_col
                current_col += 1

        # ---------- DYNAMIC COMPANY HEADER MERGE ----------
        last_data_col = current_col - 1
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=last_data_col
        )
        
        # Also merge the time interval info row if it exists
        if time_interval:
            ws.merge_cells(
                start_row=4,
                start_column=1,
                end_row=4,
                end_column=last_data_col
            )

        # ---------- GROUP DATA & CALCULATE MIN/MAX ----------
        data_map = defaultdict(dict)
        param_values = defaultdict(list)  # Store values for each parameter for min/max calculation
        
        for r in records:
            # Format timestamp based on time interval
            if time_interval:
                # For averaged data, use the interval start time
                ts = r.timestamp.strftime("%Y-%m-%d %H:%M")
                if time_interval == '1 day':
                    ts = r.timestamp.strftime("%Y-%m-%d")
                elif time_interval == '6 hours':
                    # Show both start and end of 6-hour interval
                    interval_end = r.timestamp + timedelta(hours=6)
                    ts = f"{r.timestamp.strftime('%Y-%m-%d %H:%M')} to {interval_end.strftime('%H:%M')}"
            else:
                # Original behavior: round to minute
                ts = r.timestamp.replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M")
                
            key = (r.device_id, r.parameter_id)
            data_map[ts][key] = r.value
            
            # Collect values for min/max calculation
            if r.value is not None:
                param_values[key].append(r.value)

        # ---------- WRITE TIMESTAMP DATA ----------
        current_row = data_start_row
        
        # Sort timestamps chronologically
        sorted_timestamps = sorted(data_map.keys())
        
        for ts in sorted_timestamps:
            values = data_map[ts]
            
            # Timestamp cell
            ts_cell = ws.cell(row=current_row, column=1)
            ts_cell.value = ts
            ts_cell.alignment = Alignment(horizontal="left")
            ts_cell.border = border

            # Data cells
            for key, col in column_map.items():
                cell = ws.cell(row=current_row, column=col)
                value = values.get(key)
                cell.value = value
                if value is not None:
                    # Apply number formatting if it looks like a numeric value
                    if isinstance(value, (int, float)):
                        cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                cell.border = border
                cell.alignment = center_alignment
            
            current_row += 1

        # ---------- ADD MIN/MAX ROWS ----------
        # Add some space before summary rows
        current_row += 1
        
        # Min Row
        min_row = current_row
        min_label = ws.cell(row=min_row, column=1)
        min_label.value = "Min Values:"
        min_label.font = min_max_font
        min_label.fill = min_fill
        min_label.border = border
        min_label.alignment = Alignment(horizontal="right", vertical="center")
        
        # Max Row
        max_row = current_row + 1
        max_label = ws.cell(row=max_row, column=1)
        max_label.value = "Max Values:"
        max_label.font = min_max_font
        max_label.fill = max_fill
        max_label.border = border
        max_label.alignment = Alignment(horizontal="right", vertical="center")

        # Fill min/max values for each parameter column
        for param_info in param_details:
            key = (param_info['device_id'], param_info['parameter_id'])
            values = param_values.get(key, [])
            col = param_info['column']
            
            if values:
                # Min value
                min_val = min(values)
                min_cell = ws.cell(row=min_row, column=col)
                min_cell.value = min_val
                min_cell.font = min_max_font
                min_cell.fill = min_fill
                min_cell.border = border
                min_cell.alignment = center_alignment
                if isinstance(min_val, (int, float)):
                    min_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                
                # Max value
                max_val = max(values)
                max_cell = ws.cell(row=max_row, column=col)
                max_cell.value = max_val
                max_cell.font = min_max_font
                max_cell.fill = max_fill
                max_cell.border = border
                max_cell.alignment = center_alignment
                if isinstance(max_val, (int, float)):
                    max_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            else:
                # No data for this parameter
                ws.cell(row=min_row, column=col).value = "N/A"
                ws.cell(row=max_row, column=col).value = "N/A"
                
                ws.cell(row=min_row, column=col).font = min_max_font
                ws.cell(row=max_row, column=col).font = min_max_font
                
                ws.cell(row=min_row, column=col).fill = min_fill
                ws.cell(row=max_row, column=col).fill = max_fill
                
                ws.cell(row=min_row, column=col).border = border
                ws.cell(row=max_row, column=col).border = border
                
                ws.cell(row=min_row, column=col).alignment = center_alignment
                ws.cell(row=max_row, column=col).alignment = center_alignment

        # ---------- AUTO WIDTH ----------
        for col in range(1, last_data_col + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20

        # Auto-width for timestamp column
        ws.column_dimensions['A'].width = 25

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ==========================================================
    # CSV EXPORT (UPDATED WITH TIME INTERVAL SUPPORT)
    # ==========================================================
    def _generate_csv(self, devices, company_id, start_date, end_date, time_interval=None):
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        company = Company.query.get(company_id)
        
        # Add time interval info to header if applicable
        if time_interval:
            writer.writerow([f"{company.name if company else 'Company Report'} - {time_interval} Average"])
        else:
            writer.writerow([company.name if company else "Company Report"])
        
        if time_interval:
            writer.writerow([f"Data averaged by {time_interval} intervals"])
        
        writer.writerow([])

        header = ["Timestamp"]
        column_keys = []

        for device in devices:
            parameters = Parameter.query.filter_by(device_id=device.id).all()
            for param in parameters:
                col = f"{device.name} - {param.name}"
                if param.unit:
                    col += f" ({param.unit})"

                header.append(col)
                column_keys.append((device.id, param.id))

        writer.writerow(header)

        # Fetch data based on time interval
        if time_interval:
            interval_minutes = self.TIME_INTERVALS[time_interval]
            averaged_records = self._fetch_averaged_data(devices, start_date, end_date, interval_minutes)
            
            # Convert to data map format
            data_map = defaultdict(dict)
            for record in averaged_records:
                # Format timestamp
                ts = record['interval_start'].strftime("%Y-%m-%d %H:%M")
                if time_interval == '1 day':
                    ts = record['interval_start'].strftime("%Y-%m-%d")
                elif time_interval == '6 hours':
                    interval_end = record['interval_start'] + timedelta(hours=6)
                    ts = f"{record['interval_start'].strftime('%Y-%m-%d %H:%M')} to {interval_end.strftime('%H:%M')}"
                
                data_map[ts][(record['device_id'], record['parameter_id'])] = record['avg_value']
        else:
            # Original behavior
            query = SensorData.query.filter(
                SensorData.device_id.in_([d.id for d in devices])
            )

            if start_date:
                query = query.filter(SensorData.timestamp >= start_date)
            if end_date:
                query = query.filter(SensorData.timestamp <= end_date)

            query = query.order_by(SensorData.timestamp.asc())
            records = query.all()

            data_map = defaultdict(dict)
            for r in records:
                ts = r.timestamp.replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M")
                data_map[ts][(r.device_id, r.parameter_id)] = r.value

        # Write data rows
        for ts in sorted(data_map.keys()):
            values = data_map[ts]
            row = [ts]
            for key in column_keys:
                row.append(values.get(key))
            writer.writerow(row)

        output = io.BytesIO()
        output.write(buffer.getvalue().encode("utf-8"))
        output.seek(0)
        return output